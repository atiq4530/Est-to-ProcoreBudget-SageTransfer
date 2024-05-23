####basic function with package
from openpyxl import load_workbook, workbook
import datetime
today = datetime.date.today()

import tkinter
from tkinter import *
from tkinter import Tk
from tkinter import filedialog
from tkinter import ttk
import tkinter as tk

import os
import sys
import re
import json
import csv
# import numpy

import time
import requests
currentTimeStamp= time.time()

import datetime
currentDate = datetime.datetime.now()
mm =str(currentDate.strftime("%m"))
dd=str(currentDate.strftime("%d"))
yy=str(currentDate.year)


from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter, column_index_from_string

fontStyle= ("Courier", 11, "bold")

procoreBaseUrl="https://api.procore.com"


Custom_Number_Formats=[
    "_($* #,##0.00_);_($* (#,##0.00);_($* ""-""??_);_(@_)",
    '0.00%',
]


def filePathValidationCheck(procoreTemplateFile, errorMsg):
    # or os.path.exists(LATEST_COST_CODE_File) == False
    if os.path.exists(procoreTemplateFile) == False :
        print (errorMsg)
        MessageBox("Error", errorMsg)
        sys.exit()

def removeEmptyValueRow(sheet):
   for i in range((sheet.max_row)-1, 1, -1):
        if str(sheet["C"+str(i)].value) == "" or sheet["C"+str(i)].value is None:
            sheet.delete_rows(i)
        else:
            pass


def removeProCoreEmptyValueRow(sheet, colA, colB):
   for i in range((sheet.max_row+2)-1, 1, -1):
        if (str(sheet[colA+str(i)].value) == "" or sheet[colA+str(i)].value is None) and str(sheet[colB+str(i)].value) == "" or sheet[colB+str(i)].value is None:
            sheet.delete_rows(i)
        else:
            pass


def JobCostCodeFun(heading, workbook):
    for RowNumber_JCC in range (1, 50):
        Row_JCC_RangeValue= workbook.cell(RowNumber_JCC, 1).value
        if   heading== str(Row_JCC_RangeValue):
            return RowNumber_JCC
        
def ColumnFunction(SearchHeading, rowNumber, workbook):
    mx_col=workbook.max_column
    for ColumnNumber in range (1, mx_col):
              
        Row5RangeValue =  str(workbook.cell(rowNumber, ColumnNumber).value).strip()
        if Row5RangeValue:
            if str(SearchHeading).strip() == str(Row5RangeValue).strip() :
                return (ColumnNumber)



def MessageBox(title, text):
    root = tk.Tk()
    root.withdraw()
    tkinter.messagebox.showinfo(title, text)
    root.destroy()

userInputArray=[]
def inputFromUser(title, message):
    # Top level window 
    frame = tk.Tk() 
    frame.title(title) 
    frame.geometry('400x150') 
    userInputArray.clear()
  
    def printInput(): 
        global inputValue
        userInputArray.append(inputtxt.get(1.0, "end-1c")) 
        frame.destroy()
    
    labelTitle = tk.Label(text=message)
    labelTitle.pack()
    
    # TextBox Creation 
    inputtxt = tk.Text(frame, 
                    height = 1, 
                    width = 20) 
    inputtxt.pack() 
    
    # Button Creation 
    printButton = tk.Button(frame, 
                            text = "Submit",  
                            command = printInput) 
    printButton.pack() 
    frame.mainloop() 


def WriteToExcelFile(DataArr, notMatchingData, budgetLineItems_Sheet,unMatchSheet):
    #remove all data first-- if there is old data
    if budgetLineItems_Sheet.max_row >= 2:
        for rowNumber in range(1, budgetLineItems_Sheet.max_row+1):
            for colNumber in range (1, budgetLineItems_Sheet.max_column+1):
                budgetLineItems_Sheet[get_column_letter(colNumber+1)+str(rowNumber)].value =""
             
    #Data append to budget line item sheet 
    if budgetLineItems_Sheet and DataArr:
        DataArr.insert(0, ["Cost Code","Cost Type","Description","Manual Calculation","Unit Qty","UOM","Unit Cost","Budget Amount","Start Date","End Date","Curve"])
        counter=1
        for data in DataArr:
            for col in range(0, len(data)):
                budgetLineItems_Sheet[get_column_letter(col+1)+str(counter)].value =data[col]
            counter +=1
        
        budgetLineItems_Sheet.column_dimensions['B'].width = 13
        budgetLineItems_Sheet.column_dimensions['C'].width = 22
        budgetLineItems_Sheet.column_dimensions['D'].width = 15
        
   
    ####unmatch data sheet    
    if notMatchingData:
        counter =2
        for unMatchData in (notMatchingData):
            for col in range(0, len(unMatchData)):
                unMatchSheet[get_column_letter(col+1)+str(counter)].value = unMatchData[col]
            counter +=1
            
        unMatchSheet.column_dimensions['A'].width = 15
        unMatchSheet.column_dimensions['B'].width = 40
        unMatchSheet.column_dimensions['C'].width = 40

def duplicateValueAddUp(budgetLineItems_Sheet, colA, colB, colQty, colBudget):
    # budgetAmount, unitQty= 0, 0
    
    for rowNumber5 in range(1, budgetLineItems_Sheet.max_row+1):
        costCode= str(budgetLineItems_Sheet[colA+str(rowNumber5)].value).strip()
        costType= str(budgetLineItems_Sheet[colB+str(rowNumber5)].value).strip()

        unitQty= budgetLineItems_Sheet[colQty+str(rowNumber5)].value
        budgetAmount =budgetLineItems_Sheet[colBudget+str(rowNumber5)].value
        
        if isinstance(budgetAmount, (int, float)):
            counter =0
            
            for rowNumber6 in range(1, budgetLineItems_Sheet.max_row+1):
                if costCode == str(budgetLineItems_Sheet[colA+str(rowNumber6)].value).strip() and  costType == str(budgetLineItems_Sheet[colB+str(rowNumber6)].value).strip():
                    counter +=1 ####count it has duplicate value 
                    
                    if counter >=2:
                        try:
                            currrentBudget= budgetLineItems_Sheet[colBudget+str(rowNumber6)].value 
                            currrentUnit= budgetLineItems_Sheet[colQty+str(rowNumber6)].value
                            
                            if isinstance(currrentBudget, (int, float)):
                                budgetAmount += currrentBudget
                                if isinstance(unitQty, (int,float)) and isinstance(currrentUnit, (int,float)):
                                    unitQty += currrentUnit
                                    
                                ####make row empty 
                                budgetLineItems_Sheet[colA+str(rowNumber6)].value =""
                                budgetLineItems_Sheet[colB+str(rowNumber6)].value =""
                        except:
                            pass
                      
            ####comparing budget amount increase 
            if budgetAmount  > budgetLineItems_Sheet[colBudget+str(rowNumber5)].value:
                budgetLineItems_Sheet[colQty+str(rowNumber5)].value = unitQty
                budgetLineItems_Sheet[colBudget+str(rowNumber5)] = budgetAmount
        
        
def ExcelFormat(budgetLineItems_Sheet): 
    ####apply number format
    for rowNumber9 in range(1, budgetLineItems_Sheet.max_row+1):
        budgetLineItems_Sheet["H"+str(rowNumber9)].number_format = Custom_Number_Formats[0]
    
    budgetLineItems_Sheet.column_dimensions['A'].width = 13
    budgetLineItems_Sheet.column_dimensions['B'].width = 10
    budgetLineItems_Sheet.column_dimensions['C'].width = 30
    budgetLineItems_Sheet.column_dimensions['D'].width = 15
    budgetLineItems_Sheet.column_dimensions['E'].width = 10        
    budgetLineItems_Sheet.column_dimensions['F'].width = 10        
    budgetLineItems_Sheet.column_dimensions['G'].width = 10        
    budgetLineItems_Sheet.column_dimensions['H'].width = 14     
    budgetLineItems_Sheet.column_dimensions['I'].width = 15       
    budgetLineItems_Sheet.column_dimensions['J'].width = 15       
    budgetLineItems_Sheet.column_dimensions['K'].width = 10
    ###number format